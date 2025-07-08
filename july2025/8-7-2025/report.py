import openpyxl
import os
from datetime import datetime

# === File paths ===
report_file = 'report.xlsx'
monthly_report_file = 'monthly report.xlsx'
report_filled_file = 'report_filled.xlsx'
source_files = ['mfb.xlsx', 'bosta.xlsx', 'telegraph.xlsx', 'vendors.xlsx']

# === Load the report workbook ===
report_wb = openpyxl.load_workbook(report_file)
report_ws = report_wb.active

# === Helper: Read the actual data row (row 4 in Excel) ===
def get_row3_values(file_name):
    wb = openpyxl.load_workbook(file_name, data_only=True)
    try:
        ws = wb['report']
    except KeyError:
        print(f"⚠️ Sheet 'report' not found in {file_name}. Using first sheet: {wb.sheetnames[0]}")
        ws = wb.active
    values = [cell.value for cell in ws[2][:9]]  # ✅ Row 4 (index 3), columns A to I
    return values

# === Step 1: Fill report.xlsx from source files ===
for row in report_ws.iter_rows(min_row=3, max_row=6):  # Rows 3 to 6
    file_name = row[1].value  # Column B (Source File)
    if file_name:
        file_name = file_name.strip()
        if file_name in source_files:
            try:
                values = get_row3_values(file_name)
                print(f"✅ {file_name} → {values}")
                if values:
                    for i in range(9):  # Columns C to K
                        row[i + 2].value = values[i]
            except Exception as e:
                print(f"❌ Error processing {file_name}: {e}")

# === Step 2: Add totals in row 7 ===
for col_index in range(3, 12):  # Columns C to K
    total = 0
    for row_index in range(3, 7):  # Rows 3 to 6
        value = report_ws.cell(row=row_index, column=col_index).value
        if isinstance(value, (int, float)):
            total += value
    report_ws.cell(row=7, column=col_index, value=total)

# === Save filled report ===
report_wb.save(report_filled_file)
print("✅ Saved filled report to", report_filled_file)

# === Step 3: Load or create the monthly report ===
if os.path.exists(monthly_report_file):
    monthly_wb = openpyxl.load_workbook(monthly_report_file)
    monthly_ws = monthly_wb.active
else:
    monthly_wb = openpyxl.Workbook()
    monthly_ws = monthly_wb.active
    headers = [
        'Date',
        'Out For Delivery - Number',
        'Out For Delivery - Shipping',
        'Out For Delivery - Total',
        'Completed - Number',
        'Completed - Shipping',
        'Completed - Total',
        'Returned - Number',
        'Returned - Shipping',
        'Returned - Total'
    ]
    monthly_ws.append(headers)

# === Step 4: Remove old "Running Total" rows ===
rows_to_delete = []
for i, row in enumerate(monthly_ws.iter_rows(min_row=2), start=2):
    if str(row[0].value).strip().lower() == 'running total':
        rows_to_delete.append(i)
for i in reversed(rows_to_delete):
    monthly_ws.delete_rows(i)

# === Step 5: Append today’s totals (if not already added) ===
today = datetime.today().strftime('%Y-%m-%d')
today_exists = any(str(row[0]) == today for row in monthly_ws.iter_rows(min_row=2, values_only=True))

if not today_exists:
    total_row_values = [report_ws.cell(row=7, column=col).value for col in range(3, 12)]
    monthly_ws.append([today] + total_row_values)
    print(f"✅ Added new row for today: {today}")
else:
    print(f"ℹ️ Entry for today ({today}) already exists. Skipping row append.")

# === Step 6: Add updated "Running Total" row ===
last_data_row = monthly_ws.max_row
running_total_row = last_data_row + 1
monthly_ws.cell(row=running_total_row, column=1).value = 'Running Total'

for col in range(2, 11):  # Columns B to J
    col_letter = openpyxl.utils.get_column_letter(col)
    formula = f"=SUM({col_letter}2:{col_letter}{last_data_row})"
    monthly_ws.cell(row=running_total_row, column=col).value = formula

# === Save updated monthly report ===
monthly_wb.save(monthly_report_file)
print("✅ Monthly report updated with totals.")
