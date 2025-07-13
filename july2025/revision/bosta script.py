import os
import openpyxl

# ===== Helper Function =====
def format_row(row, col_widths):
    return " | ".join(str(cell if cell is not None else "").ljust(col_widths[i]) for i, cell in enumerate(row))

# ===== Step 1: Insert Order Reference and Clean BO_Delivery =====
def process_bosta_excel():
    source_wb = openpyxl.load_workbook('bosta-completed.xlsx')
    source_ws = source_wb.active

    header = [cell.value for cell in source_ws[1]]
    if 'Order Reference' not in header:
        raise ValueError("Column 'Order Reference' not found in source sheet.")
    col_ref_index = header.index('Order Reference') + 1

    ref_values = [str(row[0].value).strip() for row in source_ws.iter_rows(min_row=2, min_col=col_ref_index, max_col=col_ref_index) if row[0].value]
    print(f"[INFO] Extracted {len(ref_values)} 'Order Reference' values.")

    target_wb = openpyxl.load_workbook('bosta.xlsx')
    delivery_ws = target_wb['BO_Delivery']

    delivery_header = [cell.value for cell in delivery_ws[1]]
    col_orders_completed = delivery_header.index('Orders completed') + 1
    col_order_number = delivery_header.index('Order Number') + 1

    # Map existing Order Numbers to their row index
    order_number_map = {}
    for row in range(2, delivery_ws.max_row + 1):
        order_num = delivery_ws.cell(row=row, column=col_order_number).value
        if order_num:
            order_number_map[str(order_num).strip()] = row

    matched = 0
    deleted = 0
    appended = 0
    next_empty_row = delivery_ws.max_row + 1
    rows_to_delete = []

    for val in ref_values:
        val_clean = str(val).strip()
        if val_clean in order_number_map:
            target_row = order_number_map[val_clean]
            delivery_ws.cell(row=target_row, column=col_orders_completed).value = val_clean
            rows_to_delete.append(target_row)
            matched += 1
        else:
            delivery_ws.cell(row=next_empty_row, column=col_orders_completed).value = val_clean
            next_empty_row += 1
            appended += 1

    # Delete matched duplicate rows (bottom-up)
    for row in sorted(rows_to_delete, reverse=True):
        delivery_ws.delete_rows(row)
        deleted += 1

    print(f"[INFO] Matched and deleted {deleted} duplicated rows from 'BO_Delivery'.")
    print(f"[INFO] Appended {appended} unmatched 'Orders completed' rows at the bottom.")

    # Append to Completed Orders sheet
    completed_ws = target_wb['Completed Orders']
    completed_header = [cell.value for cell in completed_ws[1]]
    col_start = completed_header.index('Order Status') + 1

    for row in range(completed_ws.max_row, 1, -1):
        if completed_ws.cell(row=row, column=col_start).value not in (None, ''):
            last_row = row
            break
    else:
        last_row = 1

    start_row = last_row + 1
    copied_rows = 0

    source_header = [cell.value for cell in source_ws[1]]
    if 'Order Status' not in source_header:
        raise ValueError("Column 'Order Status' not found in source sheet.")
    source_start_index = source_header.index('Order Status')

    for i, source_row in enumerate(source_ws.iter_rows(min_row=2, values_only=True), start=start_row):
        row_slice = source_row[source_start_index:]  # Start from 'Order Status'
        for j, value in enumerate(row_slice, start=col_start):
            completed_ws.cell(row=i, column=j).value = value
        copied_rows += 1

    print(f"[INFO] Appended {copied_rows} rows to 'Completed Orders' starting at row {start_row}.")

    # === Remove duplicate rows from 'Completed Orders' ===
    all_data = []
    seen_rows = set()

    for row in completed_ws.iter_rows(min_row=2, values_only=True):
        row_tuple = tuple(row)
        if any(cell is not None for cell in row):  # Skip blank rows
            if row_tuple not in seen_rows:
                seen_rows.add(row_tuple)
                all_data.append(row_tuple)

    # Clear all rows except the header
    completed_ws.delete_rows(2, completed_ws.max_row)

    # Write back only unique rows
    for i, row_data in enumerate(all_data, start=2):
        for j, value in enumerate(row_data, start=1):
            completed_ws.cell(row=i, column=j, value=value)

    print(f"[INFO] Removed duplicates. Retained {len(all_data)} unique rows in 'Completed Orders'.")

    # ===== Move RETURNED rows to Returned Orders =====
    print("[INFO] Moving RETURNED orders to 'Returned Orders' sheet...")

    returned_ws = target_wb['Returned Orders']
    returned_header = [cell.value for cell in returned_ws[1]]
    order_status_col = completed_header.index('Order Status') + 1

    # Find last used row in Returned Orders
    for row in range(returned_ws.max_row, 1, -1):
        if any(returned_ws.cell(row=row, column=col).value is not None for col in range(1, returned_ws.max_column + 1)):
            last_returned_row = row
            break
    else:
        last_returned_row = 1
    append_row = last_returned_row + 1

    # Collect RETURNED rows and their indices
    returned_rows = []
    rows_to_delete = []

    for row in range(2, completed_ws.max_row + 1):
        status = completed_ws.cell(row=row, column=order_status_col).value
        if str(status).strip().upper() == 'RETURNED':
            row_data = [completed_ws.cell(row=row, column=col).value for col in range(1, completed_ws.max_column + 1)]
            returned_rows.append(row_data)
            rows_to_delete.append(row)

    # Append to Returned Orders
    for row_data in returned_rows:
        for col_idx, value in enumerate(row_data, start=1):
            returned_ws.cell(row=append_row, column=col_idx, value=value)
        append_row += 1

    # Delete rows from Completed Orders (from bottom to top)
    for row in sorted(rows_to_delete, reverse=True):
        completed_ws.delete_rows(row)

    print(f"[INFO] Moved {len(returned_rows)} RETURNED rows to 'Returned Orders'.")

    # Save workbook
    target_wb.save('bosta.xlsx')
    print("[INFO] Workbook saved after insert/delete operations.")

# ===== MAIN =====
if __name__ == "__main__":
    print("\n=== Step 1: Process and Clean bosta.xlsx ===")
    process_bosta_excel()
