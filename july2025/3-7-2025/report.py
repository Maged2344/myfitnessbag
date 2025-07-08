import openpyxl

# File paths
report_file = 'report.xlsx'
source_files = ['mfb.xlsx', 'bosta.xlsx', 'telegraph.xlsx', 'vendors.xlsx']

# Load the report workbook
report_wb = openpyxl.load_workbook(report_file)
report_ws = report_wb.active  # Assuming it's the first sheet

# Helper: Read 3rd row values from source file
def get_row3_values(file_name):
    wb = openpyxl.load_workbook(file_name, data_only=True)

    # Use 'report' sheet or fallback to first
    try:
        ws = wb['report']
    except KeyError:
        print(f"⚠️ Sheet 'report' not found in {file_name}. Using first sheet: {wb.sheetnames[0]}")
        ws = wb.active

    values = [cell.value for cell in ws[3][:9]]  # Row 3, columns A to I
    return values

# Process each row in report.xlsx (rows 3 to 6 contain file names)
for row in report_ws.iter_rows(min_row=3, max_row=6):  # Row 3 to 6
    file_name = row[1].value  # Column B contains file name
    if file_name in source_files:
        try:
            values = get_row3_values(file_name)
            if values:
                for i in range(9):  # Columns C to K
                    row[i + 2].value = values[i]
        except Exception as e:
            print(f"❌ Error processing {file_name}: {e}")

# Add totals in row 7 (which has "Total" in column B)
total_row_index = 7
for col_index in range(3, 12):  # Columns C (3) to K (11)
    total = 0
    for row_index in range(3, 7):  # Rows 3 to 6 (data rows)
        value = report_ws.cell(row=row_index, column=col_index).value
        if isinstance(value, (int, float)):
            total += value
    report_ws.cell(row=total_row_index, column=col_index, value=total)

# Save the updated report
report_wb.save('report_filled.xlsx')
print("✅ Report updated, values copied, and totals calculated.")
