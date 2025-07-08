import openpyxl

# ===== Step 1: Insert Order Number and Clean MFB_Delivery =====
def process_mfb_excel():
    source_wb = openpyxl.load_workbook('mfb-completed.xlsx')
    source_ws = source_wb.active

    # --- Extract Order Numbers ---
    header = [cell.value for cell in source_ws[1]]
    if 'Order Number' not in header:
        raise ValueError("Column 'Order Number' not found in source sheet.")
    col_order_number = header.index('Order Number') + 1

    order_numbers = [str(row[0].value).strip() for row in source_ws.iter_rows(min_row=2, min_col=col_order_number, max_col=col_order_number) if row[0].value]
    print(f"[INFO] Extracted {len(order_numbers)} 'Order Number' values.")

    # --- Load mfb.xlsx and write order numbers ---
    target_wb = openpyxl.load_workbook('mfb.xlsx')
    delivery_ws = target_wb['MFB_Delivery']

    delivery_header = [cell.value for cell in delivery_ws[1]]
    col_orders_completed = delivery_header.index('Orders completed') + 1
    col_existing_order_number = delivery_header.index('Order Number') + 1

    for i, val in enumerate(order_numbers, start=2):
        delivery_ws.cell(row=i, column=col_orders_completed).value = val
    print(f"[INFO] Wrote {len(order_numbers)} values to 'Orders completed'.")

    # --- Remove duplicates from MFB_Delivery ---
    orders_completed_vals = {}
    existing_order_vals = {}
    max_row = delivery_ws.max_row

    for row in range(2, max_row + 1):
        oc_val = delivery_ws.cell(row=row, column=col_orders_completed).value
        on_val = delivery_ws.cell(row=row, column=col_existing_order_number).value
        if oc_val:
            orders_completed_vals[row] = str(oc_val).strip()
        if on_val:
            existing_order_vals[row] = str(on_val).strip()

    duplicated_numbers = set(orders_completed_vals.values()) & set(existing_order_vals.values())
    rows_to_delete = {row for row, val in orders_completed_vals.items() if val in duplicated_numbers}
    rows_to_delete |= {row for row, val in existing_order_vals.items() if val in duplicated_numbers}

    for row in sorted(rows_to_delete, reverse=True):
        delivery_ws.delete_rows(row)
    print(f"[INFO] Deleted {len(rows_to_delete)} rows from 'MFB_Delivery' due to duplicates.")

    # --- Append data to Completed Orders sheet ---
    completed_ws = target_wb['Completed Orders']
    completed_header = [cell.value for cell in completed_ws[1]]
    col_start = completed_header.index('Modification Date') + 1

    # Find last used row
    for row in range(completed_ws.max_row, 1, -1):
        if completed_ws.cell(row=row, column=col_start).value not in (None, ''):
            last_row = row
            break
    else:
        last_row = 1

    start_row = last_row + 1
    copied_rows = 0

    source_header = [cell.value for cell in source_ws[1]]
    if 'Modification Date' not in source_header:
        raise ValueError("Column 'Modification Date' not found in source sheet.")
    source_start_index = source_header.index('Modification Date')

    for i, source_row in enumerate(source_ws.iter_rows(min_row=2, values_only=True), start=start_row):
        row_slice = source_row[source_start_index:]
        for j, value in enumerate(row_slice, start=col_start):
            completed_ws.cell(row=i, column=j).value = value
        copied_rows += 1

    print(f"[INFO] Appended {copied_rows} rows to 'Completed Orders' starting at row {start_row}.")

    # --- Remove duplicate rows in Completed Orders ---
    all_data = []
    seen_rows = set()

    for row in completed_ws.iter_rows(min_row=2, values_only=True):
        row_tuple = tuple(row)
        if any(cell is not None for cell in row):
            if row_tuple not in seen_rows:
                seen_rows.add(row_tuple)
                all_data.append(row_tuple)

    completed_ws.delete_rows(2, completed_ws.max_row)

    for i, row_data in enumerate(all_data, start=2):
        for j, value in enumerate(row_data, start=1):
            completed_ws.cell(row=i, column=j, value=value)

    print(f"[INFO] Removed duplicates. Retained {len(all_data)} unique rows in 'Completed Orders'.")

    # === Move "MFB_Return" rows to 'Returned Orders' sheet ===
    returned_ws = target_wb['Returned Orders']

    # Get indexes of necessary columns
    completed_header = [cell.value for cell in completed_ws[1]]
    order_status_col = completed_header.index('Order Status') + 1
    payment_method_col = completed_header.index('Payment Method') + 1
    order_total_col = completed_header.index('Order Total Amount') + 1

    max_completed_row = completed_ws.max_row
    rows_to_move = []

    for row in range(2, max_completed_row + 1):
        status = completed_ws.cell(row=row, column=order_status_col).value
        if str(status).strip().lower() == 'mfb_return':
            rows_to_move.append(row)

    # Find the first empty row in Returned Orders
    last_row_returned = returned_ws.max_row
    for i in range(last_row_returned, 1, -1):
        if any(cell.value is not None for cell in returned_ws[i]):
            last_row_returned = i
            break
    else:
        last_row_returned = 1

    insert_row = last_row_returned + 1

    # Copy and delete from Completed Orders
    for row_index in sorted(rows_to_move):
        row_data = [completed_ws.cell(row=row_index, column=col).value for col in range(1, completed_ws.max_column + 1)]
        for col_index, value in enumerate(row_data, start=1):
            returned_ws.cell(row=insert_row, column=col_index).value = value
        insert_row += 1

    for row_index in sorted(rows_to_move, reverse=True):
        completed_ws.delete_rows(row_index)

    print(f"[INFO] Moved {len(rows_to_move)} 'MFB_Return' rows to 'Returned Orders'.")

    # === Zero Order Total Amount for non-cod payments ===
    for row in range(2, completed_ws.max_row + 1):
        payment = completed_ws.cell(row=row, column=payment_method_col).value
        if payment and str(payment).strip().lower() != 'cod':
            completed_ws.cell(row=row, column=order_total_col).value = 0

    print("[INFO] Set 'Order Total Amount' = 0 for non-cod payment methods.")

    # === Save workbook ===
    target_wb.save('mfb.xlsx')
    print("[INFO] Workbook saved after all operations.")


# ===== MAIN =====
if __name__ == "__main__":
    print("\n=== Step 1: Process and Clean mfb.xlsx ===")
    process_mfb_excel()
