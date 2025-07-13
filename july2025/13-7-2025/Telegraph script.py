import openpyxl

# ===== Helper Function =====
def format_row(row, col_widths):
    return " | ".join(str(cell if cell is not None else "").ljust(col_widths[i]) for i, cell in enumerate(row))

# ===== Step 1: Insert الكود and Clean TE_Delivery =====
def process_telegraph_excel():
    source_wb = openpyxl.load_workbook('telegraph-completed.xlsx')
    source_ws = source_wb.active

    header = [cell.value for cell in source_ws[1]]
    if 'الكود' not in header:
        raise ValueError("Column 'الكود' not found in source sheet.")
    col_kod_index = header.index('الكود') + 1

    kod_values = [str(row[0].value).strip() for row in source_ws.iter_rows(min_row=2, min_col=col_kod_index, max_col=col_kod_index) if row[0].value]
    print(f"[INFO] Extracted {len(kod_values)} 'الكود' values.")

    target_wb = openpyxl.load_workbook('telegraph.xlsx')
    delivery_ws = target_wb['TE_Delivery']

    delivery_header = [cell.value for cell in delivery_ws[1]]
    col_orders_completed = delivery_header.index('Orders completed') + 1
    col_order_number = delivery_header.index('Order Number') + 1

    # Map existing Order Numbers to row numbers
    order_number_map = {}
    for row in range(2, delivery_ws.max_row + 1):
        on_val = delivery_ws.cell(row=row, column=col_order_number).value
        if on_val:
            order_number_map[str(on_val).strip()] = row

    matched = 0
    deleted = 0
    appended = 0
    next_empty_row = delivery_ws.max_row + 1
    rows_to_delete = []

    for val in kod_values:
        val_clean = str(val).strip()
        if val_clean in order_number_map:
            target_row = order_number_map[val_clean]
            delivery_ws.cell(row=target_row, column=col_orders_completed).value = val_clean
            rows_to_delete.append(target_row)
            matched += 1
        else:
            delivery_ws.cell(row=next_empty_row, column=col_orders_completed).value = val_clean
            next_empty_row += 1
            appended += 1

    for row in sorted(rows_to_delete, reverse=True):
        delivery_ws.delete_rows(row)
        deleted += 1

    print(f"[INFO] Matched and deleted {deleted} duplicated rows from 'TE_Delivery'.")
    print(f"[INFO] Appended {appended} unmatched 'Orders completed' rows at the bottom.")

    # Append to Completed Orders sheet
    completed_ws = target_wb['Completed Orders']
    completed_header = [cell.value for cell in completed_ws[1]]
    col_start = completed_header.index('المستلم') + 1

    for row in range(completed_ws.max_row, 1, -1):
        if completed_ws.cell(row=row, column=col_start).value not in (None, ''):
            last_row = row
            break
    else:
        last_row = 1

    start_row = last_row + 1
    copied_rows = 0

    for i, source_row in enumerate(source_ws.iter_rows(min_row=2, values_only=True), start=start_row):
        for j, value in enumerate(source_row, start=col_start):
            completed_ws.cell(row=i, column=j).value = value
        copied_rows += 1

    print(f"[INFO] Appended {copied_rows} rows to 'Completed Orders' starting at row {start_row}.")

    target_wb.save('telegraph.xlsx')
    print("[INFO] Workbook saved after insert/delete operations.")

# ===== MAIN =====
if __name__ == "__main__":
    print("\n=== Step 1: Process and Clean telegraph.xlsx ===")
    process_telegraph_excel()
