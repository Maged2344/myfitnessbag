import os
import win32com.client
import openpyxl
from openpyxl import Workbook
from datetime import datetime

files_sheets = {
    "bosta.xlsx": ("BO_Delivery", "BO_Delivery Report"),
    "mfb.xlsx": ("MFB_Delivery", "MFB_Delivery Report"),
    "telegraph.xlsx": ("TE_Delivery", "TE_Delivery Report"),
    "vendors.xlsx": ("VE_Delivery", "VE_Delivery Report"),
}

def get_pivot_table_headers(pt):
    try:
        headers = ["Row Labels"]
        data_fields = pt.DataFields
        for i in range(1, data_fields.Count + 1):
            headers.append(data_fields.Item(i).Name)
        return headers
    except:
        return None

def update_pivot_tables():
    excel = win32com.client.DispatchEx('Excel.Application')
    excel.Visible = False

    all_report_rows = []
    grand_totals_headers = []
    grand_totals_sum = []
    max_cols = 0
    today = datetime.today()
    today_str = today.strftime("%Y-%m-%d")

    try:
        for file, (data_sheet, report_sheet) in files_sheets.items():
            full_path = os.path.abspath(file)
            if not os.path.exists(full_path):
                print(f"[ERROR] File not found: {full_path}")
                continue

            print(f"\n[INFO] Opening {file} ...")
            wb = excel.Workbooks.Open(full_path)
            try:
                ws_data = wb.Sheets(data_sheet)
                ws_report = wb.Sheets(report_sheet)
                data_address = ws_data.UsedRange.Address

                pivot_tables = ws_report.PivotTables()
                print(f"[INFO] {pivot_tables.Count} pivot table(s) in {file}")

                file_grand_total_row = None
                file_headers = []
                current_cols = 0

                for i in range(1, pivot_tables.Count + 1):
                    pt = pivot_tables.Item(i)
                    pt.ChangePivotCache(
                        wb.PivotCaches().Create(
                            SourceType=win32com.client.constants.xlDatabase,
                            SourceData=f"'{data_sheet}'!{data_address}"
                        )
                    )
                    pt.RefreshTable()

                    values = pt.TableRange2.Value
                    if not values:
                        continue

                    headers = get_pivot_table_headers(pt) or [str(c or "") for c in values[0]]
                    grand_total_row = None
                    for row in reversed(values[1:]):
                        if isinstance(row[0], str) and "grand total" in row[0].lower():
                            grand_total_row = [cell or 0 for cell in row]
                            break
                    if not grand_total_row:
                        continue

                    current_cols = max(len(headers), len(grand_total_row))
                    headers += [""] * (current_cols - len(headers))
                    grand_total_row += [0] * (current_cols - len(grand_total_row))

                    if file_grand_total_row is None:
                        file_grand_total_row = grand_total_row
                        file_headers = headers
                    else:
                        for j in range(1, current_cols):
                            try:
                                file_grand_total_row[j] += float(grand_total_row[j])
                            except:
                                pass

                if file_grand_total_row is None:
                    print(f"[WARNING] No grand total in {file}")
                    wb.Close(False)
                    continue

                if current_cols > max_cols:
                    diff = current_cols - max_cols
                    grand_totals_sum.extend([0]*diff)
                    grand_totals_headers.extend([""]*diff)
                    max_cols = current_cols

                for idx in range(current_cols):
                    if grand_totals_headers[idx] == "" and file_headers[idx]:
                        grand_totals_headers[idx] = file_headers[idx]

                all_report_rows.append([today_str, file] + file_grand_total_row)

                for j in range(1, current_cols):
                    try:
                        grand_totals_sum[j] += float(file_grand_total_row[j])
                    except:
                        pass

                wb.Save()
                wb.Close(SaveChanges=False)
                print(f"[INFO] Done with {file}")

            except Exception as e:
                print(f"[ERROR] {file}: {e}")
                wb.Close(SaveChanges=False)

    finally:
        excel.Quit()

    # --- Vendor daily totals from allvendors.xlsx ---
    vendor_file = "allvendors.xlsx"
    vendor_count = 0
    vendor_total = 0.0
    if os.path.exists(vendor_file):
        wb_v = openpyxl.load_workbook(vendor_file, data_only=True)
        ws_v = wb_v.active
        headers = [c.value for c in ws_v[1]]
        if "Order Total Amount" in headers:
            col_index = headers.index("Order Total Amount") + 1
            for cell in ws_v.iter_rows(min_row=2, min_col=col_index, max_col=col_index, values_only=True):
                val = cell[0]
                if isinstance(val, (int, float)):
                    vendor_total += val
                    vendor_count += 1
        else:
            print(f"[WARNING] 'Order Total Amount' column not found in {vendor_file}")
    else:
        print(f"[WARNING] {vendor_file} not found")

    # --- Write report.xlsx ---
    wb_r = Workbook()
    ws_r = wb_r.active
    ws_r.title = "Pivot Totals Summary"

    # Header
    headers = ["Date", "Source File"] + grand_totals_headers
    for c, h in enumerate(headers, 1):
        ws_r.cell(row=1, column=c, value=h)

    # Per-file rows
    for r, data in enumerate(all_report_rows, 2):
        for c, v in enumerate(data, 1):
            ws_r.cell(row=r, column=c, value=v)

    # Combined total
    total_row = ["", "Total"] + [vendor_total if i==1 else grand_totals_sum[i] for i in range(len(grand_totals_sum))]
    combined_row = len(all_report_rows) + 2
    for c, v in enumerate(total_row, 1):
        ws_r.cell(row=combined_row, column=c, value=v)

    # Daily vendor totals section
    vendor_header_row = combined_row + 2
    ws_r.cell(row=vendor_header_row, column=1, value="Date")
    ws_r.cell(row=vendor_header_row, column=2, value="Vendor Orders Count")
    ws_r.cell(row=vendor_header_row, column=3, value="Vendor Orders Total")
    ws_r.cell(row=vendor_header_row+1, column=1, value=today_str)
    ws_r.cell(row=vendor_header_row+1, column=2, value=vendor_count)
    ws_r.cell(row=vendor_header_row+1, column=3, value=round(vendor_total,2))

    wb_r.save("report.xlsx")
    print(f"[SUCCESS] 'report.xlsx' updated with pivot totals and daily vendor summary.")

if __name__ == "__main__":
    update_pivot_tables()
