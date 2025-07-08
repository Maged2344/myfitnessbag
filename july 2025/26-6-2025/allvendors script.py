import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import win32com.client

# === CONFIGURATION ===

vendor_map = {
    "Bosta_delivery": ("bosta.xlsx", "BO_Delivery"),
    "MFB_Delivery": ("mfb.xlsx", "MFB_Delivery"),
    "TE_Delivery": ("telegraph.xlsx", "TE_Delivery")
}

all_files = {
    "bosta.xlsx": "BO_Delivery",
    "mfb.xlsx": "MFB_Delivery",
    "telegraph.xlsx": "TE_Delivery",
    "vendors.xlsx": "VE_Delivery"
}

files_sheets = {
    "bosta.xlsx": ("BO_Delivery", "BO_Delivery Report"),
    "mfb.xlsx": ("MFB_Delivery", "MFB_Delivery Report"),
    "telegraph.xlsx": ("TE_Delivery", "TE_Delivery Report"),
    "vendors.xlsx": ("VE_Delivery", "VE_Delivery Report"),
}


# === STEP 1: PROCESS AND UPDATE VENDOR FILES ===

print("🔄 Processing and updating vendor Excel files...\n")

df = pd.read_excel("allvendors.xlsx", sheet_name="Orders")

df.columns = [
    "Modification Date", "Order Number", "Order Status", "Full Name (Billing)",
    "Phone (Billing)", "State Name (Billing)", "Payment Method",
    "Order Shipping Amount", "Order Total Amount"
]

data_columns = df.columns[df.columns.get_loc("Modification Date"):df.columns.get_loc("Order Total Amount") + 1]

df.drop_duplicates(subset=["Order Number"], inplace=True)

existing_order_numbers = set()

for file, sheet in all_files.items():
    if os.path.exists(file):
        wb = load_workbook(file)
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(min_row=2, values_only=True):
                order_no = row[1]
                if order_no:
                    existing_order_numbers.add(str(order_no).strip())

for status, group_df in df.groupby("Order Status"):
    file_name, sheet_name = vendor_map.get(status, ("vendors.xlsx", "VE_Delivery"))
    group_df = group_df[data_columns].copy()

    if file_name in ["bosta.xlsx", "mfb.xlsx", "telegraph.xlsx"]:
        group_df.loc[group_df["Payment Method"].str.lower() != "cod", "Order Total Amount"] = 0

    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            ws.append([""] + list(data_columns))
        else:
            ws = wb[sheet_name]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append([""] + list(data_columns))

    new_rows = 0
    for row in dataframe_to_rows(group_df, index=False, header=False):
        order_no = str(row[1]).strip()
        if order_no not in existing_order_numbers:
            ws.append([""] + row)
            existing_order_numbers.add(order_no)
            new_rows += 1

    data = []
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row)

    df_sheet = pd.DataFrame(data, columns=headers)
    df_sheet.drop_duplicates(subset=["Order Number"], inplace=True)

    ws.delete_rows(2, ws.max_row)
    for row in dataframe_to_rows(df_sheet, index=False, header=False):
        ws.append(row)

    wb.save(file_name)
    print(f"✅ {file_name}: {new_rows} new rows added. Duplicates removed.")

# === STEP 2: UPDATE PIVOT TABLES AND PRINT TOTALS ===

def format_row(row, col_widths):
    return " | ".join(str(cell if cell is not None else "").ljust(col_widths[i]) for i, cell in enumerate(row))

def get_pivot_table_headers(pt):
    try:
        headers = ["Row Labels"]
        data_fields = pt.DataFields
        for i in range(1, data_fields.Count + 1):
            headers.append(data_fields.Item(i).Name)
        return headers
    except Exception:
        return None

def update_pivot_tables():
    print("\n🔁 Updating pivot tables and calculating Grand Totals...\n")

    excel = win32com.client.DispatchEx('Excel.Application')
    excel.Visible = False

    grand_totals_sum = []
    grand_totals_headers = []
    max_cols = 0

    try:
        for idx_file, (file, (data_sheet, report_sheet)) in enumerate(files_sheets.items()):
            full_path = os.path.abspath(file)
            if not os.path.exists(full_path):
                print(f"File not found: {full_path}")
                continue

            print(f"\nOpening {file} ...")
            wb = excel.Workbooks.Open(full_path)
            try:
                ws_data = wb.Sheets(data_sheet)
                ws_report = wb.Sheets(report_sheet)

                data_range = ws_data.UsedRange
                data_address = data_range.Address
                print(f"Using source data: '{data_sheet}'!{data_address}")

                pivot_tables = ws_report.PivotTables()
                print(f"pivot_tables count: {pivot_tables.Count}")

                for i in range(1, pivot_tables.Count + 1):
                    pt = pivot_tables.Item(i)
                    print(f"\nUpdating pivot table '{pt.Name}' on sheet '{report_sheet}' ...")
                    pt.ChangePivotCache(
                        wb.PivotCaches().Create(
                            SourceType=win32com.client.constants.xlDatabase,
                            SourceData=f"'{data_sheet}'!{data_address}"
                        )
                    )
                    pt.RefreshTable()

                    values = pt.TableRange2.Value
                    if not values:
                        print("No data found in pivot table.")
                        continue

                    headers = get_pivot_table_headers(pt)
                    if not headers:
                        headers = [str(cell if cell is not None else "") for cell in values[0]]

                    grand_total_row = None
                    for row in reversed(values[1:] if len(values) > 1 else []):
                        if row[0] and isinstance(row[0], str) and "grand total" in row[0].lower():
                            grand_total_row = [cell if cell is not None else "" for cell in row]
                            break

                    if grand_total_row is None:
                        print("Grand Total row not found.")
                        continue

                    current_cols = max(len(headers), len(grand_total_row))
                    if current_cols > max_cols:
                        diff = current_cols - max_cols
                        for _ in range(diff):
                            grand_totals_sum.append(0.0)
                            grand_totals_headers.append("")
                        max_cols = current_cols

                    headers += [""] * (max_cols - len(headers))
                    grand_total_row += [0] * (max_cols - len(grand_total_row))

                    for idx_col in range(max_cols):
                        if grand_totals_headers[idx_col] == "" and headers[idx_col] != "":
                            grand_totals_headers[idx_col] = headers[idx_col]

                    col_widths = [max(len(str(grand_totals_headers[i])), len(str(grand_total_row[i]))) for i in range(max_cols)]
                    print(format_row(grand_totals_headers, col_widths))
                    print(format_row(grand_total_row, col_widths))

                    for idx_col in range(1, max_cols):
                        try:
                            val = float(grand_total_row[idx_col])
                            grand_totals_sum[idx_col] += val
                        except (TypeError, ValueError):
                            pass

                wb.Save()
                print(f"Updated and saved {file}")

                print("\n" + "=" * 80 + "\n")
                if idx_file != len(files_sheets) - 1:
                    print("-" * 80 + "\n")

            except Exception as e:
                print(f"Error processing {file}: {e}")
            finally:
                wb.Close(SaveChanges=False)

    finally:
        excel.Quit()

    if grand_totals_sum and grand_totals_headers:
        print("\n=== Total sums of Grand Totals across all pivot tables ===\n")
        col_widths = [max(len(str(grand_totals_headers[i])), len(str(int(grand_totals_sum[i])) if grand_totals_sum[i].is_integer() else f"{grand_totals_sum[i]:.2f}")) for i in range(max_cols)]
        print(format_row(grand_totals_headers, col_widths))

        total_sum_row = ["Grand Total"]
        for idx_col in range(1, max_cols):
            val = grand_totals_sum[idx_col]
            total_sum_row.append(str(int(val)) if val.is_integer() else f"{val:.2f}")
        total_sum_row += [""] * (max_cols - len(total_sum_row))
        print(format_row(total_sum_row, col_widths))

    print("\n✅ All pivot tables updated!")

# === MAIN ===
if __name__ == "__main__":
    update_pivot_tables()
