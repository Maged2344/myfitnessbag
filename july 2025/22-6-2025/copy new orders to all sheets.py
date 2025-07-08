import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

# Mapping of Order Status to file and sheet
vendor_map = {
    "Bosta_delivery": ("bosta.xlsx", "BO_Delivery"),
    "MFB_Delivery": ("mfb.xlsx", "MFB_Delivery"),
    "TE_Delivery": ("telegraph.xlsx", "TE_Delivery")
}

# All output files with their main sheet names
all_files = {
    "bosta.xlsx": "BO_Delivery",
    "mfb.xlsx": "MFB_Delivery",
    "telegraph.xlsx": "TE_Delivery",
    "vendors.xlsx": "VE_Delivery"
}

# Load the source data
df = pd.read_excel("allvendors.xlsx", sheet_name="Orders")

# Rename columns
df.columns = [
    "Modification Date", "Order Number", "Order Status", "Full Name (Billing)",
    "Phone (Billing)", "State Name (Billing)", "Payment Method",
    "Order Shipping Amount", "Order Total Amount"
]

# Keep only needed columns (B to I in Excel terms, index 1 to 8)
data_columns = df.columns[df.columns.get_loc("Modification Date"):df.columns.get_loc("Order Total Amount") + 1]

# ✅ Remove duplicates from source based on Order Number
df.drop_duplicates(subset=["Order Number"], inplace=True)

# ✅ Gather all existing order numbers globally (across all files)
existing_order_numbers = set()

for file, sheet in all_files.items():
    if os.path.exists(file):
        wb = load_workbook(file)
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(min_row=2, values_only=True):
                order_no = row[1]
                if order_no:
                    existing_order_numbers.add(str(order_no).strip())

# ✅ Group and route to correct files/sheets
for status, group_df in df.groupby("Order Status"):
    file_name, sheet_name = vendor_map.get(status, ("vendors.xlsx", "VE_Delivery"))
    group_df = group_df[data_columns].copy()

    # If not COD, set Order Total Amount to 0 (for Bosta, MFB, Telegraph only)
    if file_name in ["bosta.xlsx", "mfb.xlsx", "telegraph.xlsx"]:
        group_df.loc[group_df["Payment Method"].str.lower() != "cod", "Order Total Amount"] = 0

    # Load or create workbook
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            ws.append([""] + list(data_columns))  # header with blank col A
        else:
            ws = wb[sheet_name]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append([""] + list(data_columns))  # header with blank col A

    # Append new rows (skip if Order Number exists)
    new_rows = 0
    for row in dataframe_to_rows(group_df, index=False, header=False):
        order_no = str(row[1]).strip()
        if order_no not in existing_order_numbers:
            ws.append([""] + row)
            existing_order_numbers.add(order_no)
            new_rows += 1

    # ✅ After writing, deduplicate inside the sheet again just in case
    data = []
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row)

    df_sheet = pd.DataFrame(data, columns=headers)
    df_sheet.drop_duplicates(subset=["Order Number"], inplace=True)

    # Clear the sheet and re-write header and unique data
    ws.delete_rows(2, ws.max_row)
    for row in dataframe_to_rows(df_sheet, index=False, header=False):
        ws.append(row)

    wb.save(file_name)
    print(f"✅ {file_name}: {new_rows} new rows added. Duplicates removed.")

print("\n🎉 Done! All vendor sheets updated and cleaned with no duplicated Order Numbers.")
