import os
import openpyxl
import win32com.client

# ===== Helper Functions =====
def format_row(row, col_widths):
    return " | ".join(str(cell if cell is not None else "").ljust(col_widths[i]) for i, cell in enumerate(row))

def get_pivot_table_headers(pt):
    try:
        headers = ["Row Labels"]
        data_fields = pt.DataFields
        for i in range(1, data_fields.Count + 1):
            headers.append(data_fields.Item(i).Name)
        return headers
    except Exception:
        return None

# ===== Step 1: Insert الكود and Clean TE_Delivery =====
def process_telegraph_excel():
    source_wb = openpyxl.load_workbook('telegraph-completed.xlsx')
    source_ws = source_wb.active

    header = [cell.value for cell in source_ws[1]]
    if 'الكود' not in header:
        raise ValueError("Column 'الكود' not found in source sheet.")
    col_kod_index = header.index('الكود') + 1

    kod_values = [str(row[0].value).strip() for row in source_ws.iter_rows(min_row=2, min_col=col_kod_index, max_col=col_kod_index) if row[0].value]
    print(f"[INFO] Extracted {len(kod_values)} 'الكود' values.")

    target_wb = openpyxl.load_workbook('telegraph.xlsx')
    delivery_ws = target_wb['TE_Delivery']

    delivery_header = [cell.value for cell in delivery_ws[1]]
    col_orders_completed = delivery_header.index('Orders completed') + 1
    col_order_number = delivery_header.index('Order Number') + 1

    for i, val in enumerate(kod_values, start=2):
        delivery_ws.cell(row=i, column=col_orders_completed).value = val
    print(f"[INFO] Wrote {len(kod_values)} values to 'Orders completed'.")

    orders_completed_vals = {}
    order_number_vals = {}
    max_row = delivery_ws.max_row

    for row in range(2, max_row + 1):
        oc_val = delivery_ws.cell(row=row, column=col_orders_completed).value
        on_val = delivery_ws.cell(row=row, column=col_order_number).value
        if oc_val:
            orders_completed_vals[row] = str(oc_val).strip()
        if on_val:
            order_number_vals[row] = str(on_val).strip()

    duplicated_numbers = set(orders_completed_vals.values()) & set(order_number_vals.values())
    rows_to_delete = {row for row, val in orders_completed_vals.items() if val in duplicated_numbers}
    rows_to_delete |= {row for row, val in order_number_vals.items() if val in duplicated_numbers}

    for row in sorted(rows_to_delete, reverse=True):
        delivery_ws.delete_rows(row)
    print(f"[INFO] Deleted {len(rows_to_delete)} rows from 'TE_Delivery' due to duplicates.")

    # Append to Completed Orders sheet
    completed_ws = target_wb['Completed Orders']
    completed_header = [cell.value for cell in completed_ws[1]]
    col_start = completed_header.index('المستلم') + 1

    for row in range(completed_ws.max_row, 1, -1):
        if completed_ws.cell(row=row, column=col_start).value not in (None, ''):
            last_row = row
            break
    else:
        last_row = 1

    start_row = last_row + 1
    copied_rows = 0

    for i, source_row in enumerate(source_ws.iter_rows(min_row=2, values_only=True), start=start_row):
        for j, value in enumerate(source_row, start=col_start):
            completed_ws.cell(row=i, column=j).value = value
        copied_rows += 1

    print(f"[INFO] Appended {copied_rows} rows to 'Completed Orders' starting at row {start_row}.")

    target_wb.save('telegraph.xlsx')
    print("[INFO] Workbook saved after insert/delete operations.")

# ===== Step 2: Refresh Pivot Tables in TE_Delivery Report =====
def update_telegraph_pivot():
    file = "telegraph.xlsx"
    data_sheet = "TE_Delivery"
    report_sheet = "TE_Delivery Report"

    excel = win32com.client.DispatchEx('Excel.Application')
    excel.Visible = False

    try:
        full_path = os.path.abspath(file)
        if not os.path.exists(full_path):
            print(f"[ERROR] File not found: {full_path}")
            return

        wb = excel.Workbooks.Open(full_path)
        ws_data = wb.Sheets(data_sheet)
        ws_report = wb.Sheets(report_sheet)

        data_range = ws_data.UsedRange
        data_address = data_range.Address
        print(f"[INFO] Updating pivot using data range: '{data_sheet}'!{data_address}")

        pivot_tables = ws_report.PivotTables()
        print(f"[INFO] Found {pivot_tables.Count} pivot table(s) in '{report_sheet}'.")

        grand_totals_sum = []
        grand_totals_headers = []
        max_cols = 0

        for i in range(1, pivot_tables.Count + 1):
            pt = pivot_tables.Item(i)
            print(f"\n[INFO] Refreshing pivot table '{pt.Name}'...")
            pt.ChangePivotCache(
                wb.PivotCaches().Create(
                    SourceType=win32com.client.constants.xlDatabase,
                    SourceData=f"'{data_sheet}'!{data_address}"
                )
            )
            pt.RefreshTable()

            values = pt.TableRange2.Value
            if not values:
                print("[WARNING] Pivot table is empty.")
                continue

            headers = get_pivot_table_headers(pt)
            if not headers:
                headers = list(values[0]) if isinstance(values[0], tuple) else [values[0]]

            grand_total_row = None
            for row in reversed(values[1:]):
                if row and isinstance(row[0], str) and "grand total" in row[0].lower():
                    grand_total_row = [cell if cell is not None else "" for cell in row]
                    break

            if grand_total_row is None:
                print("[WARNING] Grand Total row not found.")
                continue

            current_cols = max(len(headers), len(grand_total_row))
            if current_cols > max_cols:
                for _ in range(current_cols - max_cols):
                    grand_totals_sum.append(0.0)
                    grand_totals_headers.append("")
                max_cols = current_cols

            headers += [""] * (max_cols - len(headers))
            grand_total_row += [0] * (max_cols - len(grand_total_row))

            for idx_col in range(max_cols):
                if grand_totals_headers[idx_col] == "" and headers[idx_col] != "":
                    grand_totals_headers[idx_col] = headers[idx_col]

            col_widths = [max(len(str(grand_totals_headers[i])), len(str(grand_total_row[i]))) for i in range(max_cols)]

            print(format_row(grand_totals_headers, col_widths))
            print(format_row(grand_total_row, col_widths))

            for idx_col in range(1, max_cols):
                try:
                    val = float(grand_total_row[idx_col])
                    grand_totals_sum[idx_col] += val
                except (TypeError, ValueError):
                    pass

        wb.Save()
        wb.Close(SaveChanges=False)
        print(f"[INFO] Pivot table(s) updated and '{file}' saved.")

        # Final summary
        print("\n=== Total Grand Totals Summary ===\n")
        col_widths = [max(len(str(grand_totals_headers[i])), len(str(int(grand_totals_sum[i])) if grand_totals_sum[i].is_integer() else f"{grand_totals_sum[i]:.2f}")) for i in range(max_cols)]
        print(format_row(grand_totals_headers, col_widths))

        total_row = ["Grand Total"]
        for idx_col in range(1, max_cols):
            val = grand_totals_sum[idx_col]
            total_row.append(str(int(val)) if val.is_integer() else f"{val:.2f}")
        total_row += [""] * (max_cols - len(total_row))
        print(format_row(total_row, col_widths))

    finally:
        excel.Quit()
        print("\n[SUCCESS] All pivot tables refreshed and closed Excel instance.")

# ===== MAIN =====
if __name__ == "__main__":
    print("\n=== Step 1: Process and Clean telegraph.xlsx ===")
    process_telegraph_excel()

    print("\n=== Step 2: Update Pivot Tables in telegraph.xlsx ===")
    update_telegraph_pivot()
