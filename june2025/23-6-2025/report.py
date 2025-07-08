import os
import win32com.client
import openpyxl
from openpyxl import Workbook
from datetime import datetime

files_sheets = {
    "bosta.xlsx": ("BO_Delivery", "BO_Delivery Report"),
    "mfb.xlsx": ("MFB_Delivery", "MFB_Delivery Report"),
    "telegraph.xlsx": ("TE_Delivery", "TE_Delivery Report"),
    "vendors.xlsx": ("VE_Delivery", "VE_Delivery Report"),
}

def format_row(row, col_widths):
    return " | ".join(str(cell if cell is not None else "").ljust(col_widths[i]) for i, cell in enumerate(row))

def get_pivot_table_headers(pt):
    try:
        headers = ["Row Labels"]
        data_fields = pt.DataFields
        for i in range(1, data_fields.Count + 1):
            headers.append(data_fields.Item(i).Name)
        return headers
    except Exception:
        return None

def update_pivot_tables():
    excel = win32com.client.DispatchEx('Excel.Application')
    excel.Visible = False

    all_report_rows = []
    grand_totals_headers = []
    grand_totals_sum = []
    max_cols = 0
    today_str = datetime.today().strftime("%Y-%m-%d")

    try:
        for file, (data_sheet, report_sheet) in files_sheets.items():
            full_path = os.path.abspath(file)
            if not os.path.exists(full_path):
                print(f"[ERROR] File not found: {full_path}")
                continue

            print(f"\n[INFO] Opening {file} ...")
            wb = excel.Workbooks.Open(full_path)

            try:
                ws_data = wb.Sheets(data_sheet)
                ws_report = wb.Sheets(report_sheet)

                data_range = ws_data.UsedRange
                data_address = data_range.Address

                pivot_tables = ws_report.PivotTables()
                print(f"[INFO] {pivot_tables.Count} pivot table(s) found in {file}")

                file_grand_total_row = None
                file_headers = []
                current_cols = 0

                for i in range(1, pivot_tables.Count + 1):
                    pt = pivot_tables.Item(i)
                    print(f"[INFO] Refreshing pivot table: {pt.Name}")
                    pt.ChangePivotCache(
                        wb.PivotCaches().Create(
                            SourceType=win32com.client.constants.xlDatabase,
                            SourceData=f"'{data_sheet}'!{data_address}"
                        )
                    )
                    pt.RefreshTable()

                    values = pt.TableRange2.Value
                    if not values:
                        print("Pivot table is empty.")
                        continue

                    headers = get_pivot_table_headers(pt)
                    if not headers:
                        headers = [str(cell if cell else "") for cell in values[0]]

                    grand_total_row = None
                    for row in reversed(values[1:] if len(values) > 1 else []):
                        if isinstance(row[0], str) and "grand total" in row[0].lower():
                            grand_total_row = [cell if cell is not None else 0 for cell in row]
                            break

                    if not grand_total_row:
                        continue

                    current_cols = max(len(headers), len(grand_total_row))
                    if len(grand_total_row) < current_cols:
                        grand_total_row += [0] * (current_cols - len(grand_total_row))
                    if len(headers) < current_cols:
                        headers += [""] * (current_cols - len(headers))

                    if not file_grand_total_row:
                        file_grand_total_row = grand_total_row
                        file_headers = headers
                    else:
                        for j in range(1, current_cols):
                            try:
                                file_grand_total_row[j] += float(grand_total_row[j])
                            except:
                                pass

                if not file_grand_total_row:
                    print(f"[WARNING] No grand total row found in {file}")
                    continue

                # Initialize total headers and master sum
                if current_cols > max_cols:
                    for _ in range(current_cols - max_cols):
                        grand_totals_sum.append(0.0)
                        grand_totals_headers.append("")
                    max_cols = current_cols

                for idx in range(current_cols):
                    if grand_totals_headers[idx] == "" and file_headers[idx] != "":
                        grand_totals_headers[idx] = file_headers[idx]

                # Add to all_report_rows
                report_row = [today_str, file] + file_grand_total_row
                all_report_rows.append(report_row)

                for i in range(1, current_cols):
                    try:
                        val = float(file_grand_total_row[i])
                        grand_totals_sum[i] += val
                    except:
                        pass

                wb.Save()
                print(f"[INFO] Saved and closed {file}.")

            finally:
                wb.Close(SaveChanges=False)

    finally:
        excel.Quit()

    # Write to report.xlsx
    if all_report_rows:
        print("\n[INFO] Writing to 'report.xlsx' ...")
        wb_report = Workbook()
        ws = wb_report.active
        ws.title = "Pivot Totals Summary"

        # Header
        headers = ["Date", "Source File"] + grand_totals_headers
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # Data Rows
        for row_idx, row_data in enumerate(all_report_rows, start=2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Final Total Row
        total_row = ["", "Total"] + [
            int(val) if isinstance(val, float) and val.is_integer() else round(val, 2)
            for val in grand_totals_sum
        ]
        for col_idx, value in enumerate(total_row, 1):
            ws.cell(row=len(all_report_rows) + 2, column=col_idx, value=value)

        wb_report.save("report.xlsx")
        print("[SUCCESS] Report written to 'report.xlsx'.")

    else:
        print("[WARNING] No data to write to report.xlsx.")

if __name__ == "__main__":
    update_pivot_tables()
